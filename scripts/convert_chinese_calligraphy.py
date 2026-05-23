from __future__ import annotations

import argparse
import json
import re
from html.parser import HTMLParser
from pathlib import Path
from typing import Any

import openpyxl


TOPIC_ID = "chinese-calligraphy-history"

DYNASTIES: list[dict[str, Any]] = [
    {"id": "legend", "nameEn": "Legendary Age", "nameCn": "传说时代", "start": -3000, "end": -1601, "color": "#6f4e37", "bg": "#faf5ed"},
    {"id": "shang", "nameEn": "Shang", "nameCn": "商", "start": -1600, "end": -1046, "color": "#8a5a2b", "bg": "#fbf3e6"},
    {"id": "western_zhou", "nameEn": "Western Zhou", "nameCn": "西周", "start": -1046, "end": -771, "color": "#a97142", "bg": "#fbf3e6"},
    {"id": "spring_autumn", "nameEn": "Spring and Autumn", "nameCn": "春秋", "start": -770, "end": -476, "color": "#94613d", "bg": "#fbf0df"},
    {"id": "warring_states", "nameEn": "Warring States", "nameCn": "战国", "start": -475, "end": -221, "color": "#7c5a3a", "bg": "#fbf0df"},
    {"id": "qin", "nameEn": "Qin", "nameCn": "秦", "start": -221, "end": -206, "color": "#3f5f4a", "bg": "#eef6ef"},
    {"id": "western_han", "nameEn": "Western Han", "nameCn": "西汉", "start": -202, "end": 8, "color": "#517a52", "bg": "#eef6ef"},
    {"id": "eastern_han", "nameEn": "Eastern Han", "nameCn": "东汉", "start": 25, "end": 220, "color": "#6b8f5f", "bg": "#eef6ef"},
    {"id": "three_kingdoms", "nameEn": "Three Kingdoms", "nameCn": "三国", "start": 220, "end": 280, "color": "#3d6f8f", "bg": "#edf4f7"},
    {"id": "western_jin", "nameEn": "Western Jin", "nameCn": "西晋", "start": 265, "end": 317, "color": "#4f8490", "bg": "#edf4f7"},
    {"id": "eastern_jin", "nameEn": "Eastern Jin", "nameCn": "东晋", "start": 317, "end": 420, "color": "#5d789b", "bg": "#edf4f7"},
    {"id": "northern_southern", "nameEn": "Northern and Southern Dynasties", "nameCn": "南北朝", "start": 420, "end": 589, "color": "#516ba8", "bg": "#eef1fb"},
    {"id": "sui", "nameEn": "Sui", "nameCn": "隋", "start": 581, "end": 618, "color": "#5b4b8a", "bg": "#f0eef8"},
    {"id": "tang", "nameEn": "Tang", "nameCn": "唐", "start": 618, "end": 907, "color": "#7a4f9a", "bg": "#f3edf8"},
    {"id": "five_dynasties", "nameEn": "Five Dynasties", "nameCn": "五代", "start": 907, "end": 960, "color": "#7b5ab6", "bg": "#f3edf8"},
    {"id": "northern_song", "nameEn": "Northern Song", "nameCn": "北宋", "start": 960, "end": 1127, "color": "#8b5a83", "bg": "#fbf0f7"},
    {"id": "southern_song", "nameEn": "Southern Song", "nameCn": "南宋", "start": 1127, "end": 1279, "color": "#9a5f8f", "bg": "#fbf0f7"},
    {"id": "yuan", "nameEn": "Yuan", "nameCn": "元", "start": 1271, "end": 1368, "color": "#7e4f9f", "bg": "#f7effb"},
    {"id": "ming", "nameEn": "Ming", "nameCn": "明", "start": 1368, "end": 1644, "color": "#9e2f2f", "bg": "#fbefed"},
    {"id": "qing", "nameEn": "Qing", "nameCn": "清", "start": 1644, "end": 1912, "color": "#b24a3a", "bg": "#fbefed"},
    {"id": "modern_early", "nameEn": "Early Modern China", "nameCn": "近代", "start": 1840, "end": 1949, "color": "#b65b48", "bg": "#fbf1ec"},
    {"id": "modern", "nameEn": "Modern China", "nameCn": "现代", "start": 1949, "end": 2026, "color": "#be6b51", "bg": "#fbf1ec"},
]

DYNASTY_BY_NAME = {item["nameCn"]: item for item in DYNASTIES}

STYLE_RULES = [
    ("oracle_bone", "甲骨文 / Oracle-bone Script", ("甲骨", "卜辞")),
    ("bronze_script", "金文 / Bronze Script", ("金文", "钟鼎", "铭文", "毛公鼎", "散氏盘")),
    ("seal_script", "篆书 / Seal Script", ("篆", "石鼓文", "峄山碑", "小篆")),
    ("clerical_script", "隶书与汉碑 / Clerical Script", ("隶", "汉碑", "张迁碑", "礼器碑", "曹全碑", "简牍")),
    ("regular_script", "楷书 / Regular Script", ("楷", "墓志", "碑", "欧阳询", "颜真卿", "柳公权", "褚遂良")),
    ("running_script", "行书 / Running Script", ("行书", "兰亭", "祭侄文稿", "苏轼", "黄庭坚", "米芾", "赵孟頫")),
    ("cursive_script", "草书 / Cursive Script", ("草书", "狂草", "怀素", "张旭", "自叙帖")),
    ("seal_carving", "篆刻与印学 / Seal Carving", ("篆刻", "印谱", "印章")),
    ("theory_collection", "书论与收藏 / Theory and Connoisseurship", ("书论", "书谱", "法书", "墨缘", "著", "撰", "刻帖", "汇帖")),
    ("exhibition_education", "展览与教育 / Exhibition and Education", ("展览", "书法家协会", "学院", "美术馆")),
]

IMPORTANT_TERMS = (
    "仓颉", "甲骨文", "石鼓文", "峄山碑", "说文解字", "兰亭序", "祭侄文稿", "自叙帖",
    "欧阳询", "颜真卿", "柳公权", "褚遂良", "张旭", "怀素", "苏轼", "黄庭坚", "米芾",
    "赵孟頫", "董其昌", "王铎", "康有为", "碑学", "书法家协会",
)


class TextExtractor(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.parts: list[str] = []

    def handle_data(self, data: str) -> None:
        text = data.strip()
        if text:
            self.parts.append(text)


def clean_html(value: Any) -> str | None:
    if not value:
        return None
    parser = TextExtractor()
    parser.feed(str(value))
    text = " ".join(parser.parts)
    text = re.sub(r"\s+", " ", text).strip()
    return text or None


def extract_first_image(value: Any) -> str | None:
    if not value:
        return None
    match = re.search(r"<img[^>]+src=['\"]([^'\"]+)['\"]", str(value), re.I)
    return match.group(1) if match else None


def parse_fuzzy_date(value: Any, fallback_year: int) -> dict[str, Any]:
    if not value:
        return {"year": fallback_year, "accuracy": "approximate"}

    raw = str(value).strip()
    is_bc = "公元前" in raw or re.search(r"(^|[^\u4e00-\u9fff])前\s*\d+", raw) is not None
    accuracy = "exact"
    if "约" in raw or "前后" in raw or "左右" in raw:
        accuracy = "approximate"
    if "以前" in raw:
        accuracy = "not_after"
    elif "以后" in raw:
        accuracy = "not_before"

    century = re.search(r"(\d+)\s*世纪", raw)
    if century:
        number = int(century.group(1))
        year = (number - 1) * 100 + 50
        return {"year": -year if is_bc else year, "accuracy": accuracy}

    year_match = re.search(r"(\d{1,4})\s*年", raw)
    if not year_match:
        year_match = re.search(r"(\d{1,4})", raw)
    if not year_match:
        return {"year": fallback_year, "accuracy": "approximate"}

    year = int(year_match.group(1))
    result: dict[str, Any] = {"year": -year if is_bc else year}

    month_match = re.search(r"年\s*(\d{1,2})\s*月", raw)
    day_match = re.search(r"月\s*(\d{1,2})\s*日", raw)
    if month_match:
        result["month"] = int(month_match.group(1))
    if day_match:
        result["day"] = int(day_match.group(1))
    if accuracy != "exact":
        result["accuracy"] = accuracy
    return result


def classify_style(title: str, description: str | None) -> str:
    haystack = f"{title} {description or ''}"
    for style_id, label, keywords in STYLE_RULES:
        if any(keyword in haystack for keyword in keywords):
            return label
    return "书法事件 / Calligraphy Event"


def importance_for(title: str, image_url: str | None, row_number: int) -> int:
    if any(term in title for term in IMPORTANT_TERMS):
        return 1
    if row_number <= 120 or row_number % 120 == 0:
        return 2
    return 4 if image_url else 5


def make_periods() -> list[dict[str, Any]]:
    return [
        {
            "id": item["id"],
            "nameEn": item["nameEn"],
            "nameCn": item["nameCn"],
            "start": {"year": item["start"]},
            "end": {"year": item["end"]},
            "colorHex": item["color"],
            "colorBackground": item["bg"],
        }
        for item in DYNASTIES
    ]


def make_streams() -> list[dict[str, Any]]:
    return [
        {
            "id": f"s_{item['id']}",
            "periodId": item["id"],
            "nameEn": item["nameEn"],
            "nameCn": item["nameCn"],
            "start": {"year": item["start"]},
            "end": {"year": item["end"]},
            "color": item["color"],
            "lane": index % 4,
            "descriptionCn": f"{item['nameCn']}书法纪年条目，按作品、书家、碑帖、书论与展览事件组织。",
        }
        for index, item in enumerate(DYNASTIES)
    ]


def convert(input_path: Path, output_root: Path) -> None:
    workbook = openpyxl.load_workbook(input_path, data_only=True)
    sheet = workbook.active
    headers = [sheet.cell(1, col).value for col in range(1, sheet.max_column + 1)]
    col = {name: index + 1 for index, name in enumerate(headers)}

    events: list[dict[str, Any]] = []
    for row in range(2, sheet.max_row + 1):
        dynasty_name = sheet.cell(row, col["朝代"]).value or "近代"
        dynasty = DYNASTY_BY_NAME.get(str(dynasty_name), DYNASTY_BY_NAME["近代"])
        raw_description = sheet.cell(row, col["事件描述"]).value
        description = clean_html(raw_description)
        title = str(sheet.cell(row, col["事件标题"]).value or "").strip()
        image_url = sheet.cell(row, col["事件图像"]).value or extract_first_image(raw_description)
        start = parse_fuzzy_date(sheet.cell(row, col["公元纪年起始"]).value, dynasty["start"])
        end_raw = sheet.cell(row, col["公元纪年截止"]).value
        end = parse_fuzzy_date(end_raw, start["year"]) if end_raw else None
        creator = sheet.cell(row, col["事件人物姓名"]).value
        traditional_date = sheet.cell(row, col["传统纪年"]).value

        event: dict[str, Any] = {
            "id": str(sheet.cell(row, col["Id"]).value or f"calligraphy-{row - 1}"),
            "periodId": dynasty["id"],
            "streamId": f"s_{dynasty['id']}",
            "titleEn": title,
            "titleCn": title,
            "date": start,
            "importance": importance_for(title, image_url, row - 1),
            "tags": [dynasty["nameCn"], classify_style(title, description)],
            "meta": {
                "dynasty": dynasty["nameCn"],
                "traditionalDate": traditional_date or "",
                "person": creator or "",
                "style": classify_style(title, description),
                "sourceOrder": row - 1,
            },
        }
        if end and end["year"] != start["year"]:
            event["endDate"] = end
        if creator:
            event["creatorCn"] = str(creator)
        if image_url:
            event["imageUrl"] = str(image_url)
        if description:
            event["descriptionCn"] = description
        events.append(event)

    events.sort(key=lambda item: (item["date"]["year"], item["meta"]["sourceOrder"]))

    topic_dir = output_root / TOPIC_ID
    chunks_dir = topic_dir / "chunks"
    chunks_dir.mkdir(parents=True, exist_ok=True)

    topic = {
        "id": TOPIC_ID,
        "titleEn": "Chinese Calligraphy History",
        "titleCn": "中国书法史",
        "description": "由《中国书法纪年表》转换而来的主题展，按朝代呈现书体、书家、碑帖、书论与展览事件。",
        "color": "#8C3A2B",
        "defaultViewport": {"startYear": -1600, "endYear": 2026},
        "minZoomRange": 1,
        "maxZoomRange": 8000,
        "playbackSpeed": 1.4,
        "chunked": True,
        "eventFields": [
            {"key": "dynasty", "labelEn": "Dynasty", "labelCn": "朝代", "type": "text"},
            {"key": "traditionalDate", "labelEn": "Traditional Date", "labelCn": "传统纪年", "type": "text"},
            {"key": "person", "labelEn": "Person", "labelCn": "人物", "type": "text"},
            {"key": "style", "labelEn": "Category", "labelCn": "书法类别", "type": "text"},
        ],
    }

    write_json(topic_dir / "topic.json", topic)
    write_json(topic_dir / "periods.json", make_periods())
    write_json(topic_dir / "streams.json", make_streams())
    write_json(topic_dir / "events.json", [])

    chunk_ranges = [
        (-3000, -501),
        (-500, -1),
        (0, 399),
        (400, 599),
        (600, 899),
        (900, 1199),
        (1200, 1399),
        (1400, 1599),
        (1600, 1799),
        (1800, 1899),
        (1900, 1949),
        (1950, 2026),
    ]
    manifest_chunks = []
    for start, end in chunk_ranges:
        chunk_events = [event for event in events if start <= event["date"]["year"] <= end]
        if not chunk_events:
            continue
        filename = f"chunks/events_chunk_{start}_{end}.json"
        write_json(topic_dir / filename, chunk_events)
        manifest_chunks.append({"start": start, "end": end, "file": filename, "count": len(chunk_events)})

    write_json(topic_dir / "manifest.json", {"topicId": TOPIC_ID, "chunks": manifest_chunks})
    update_index(output_root / "index.json")

    image_count = sum(1 for event in events if event.get("imageUrl"))
    print(f"Converted {len(events)} events, {image_count} with images, {len(manifest_chunks)} chunks.")


def update_index(index_path: Path) -> None:
    topics = json.loads(index_path.read_text(encoding="utf-8"))
    entry = {
        "id": TOPIC_ID,
        "titleCn": "中国书法史",
        "titleEn": "Chinese Calligraphy History",
        "color": "#8C3A2B",
    }
    topics = [topic for topic in topics if topic.get("id") != TOPIC_ID]
    topics.append(entry)
    write_json(index_path, topics)


def write_json(path: Path, data: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("input", type=Path)
    parser.add_argument("--output", type=Path, default=Path("public/data"))
    args = parser.parse_args()
    convert(args.input, args.output)


if __name__ == "__main__":
    main()
